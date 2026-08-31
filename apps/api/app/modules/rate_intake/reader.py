"""Reading a filled-in intake sheet, from .xlsx or .csv, into plain dicts.

One row per price, columns named by the template's header. The reader's only job
is getting cells out faithfully; every judgement about what a value *means* lives
in :mod:`.normalise`, and every judgement about whether a row can be stored lives
in :mod:`.service`.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.errors import AppError
from app.modules.rate_intake.normalise import clean

# The template's columns. A sheet missing one of these is refused up front rather
# than importing a column short and blaming the data.
COLUMNS: tuple[str, ...] = (
    "row_type",
    "property_name",
    "destination",
    "room_type",
    "room_sleeps",
    "meal_plan",
    "guest_residence",
    "price_covers",
    "label",
    "valid_from",
    "valid_to",
    "currency",
    "amount",
    "charged_per",
    "rack_or_sto",
    "discount_percent",
    "vat",
    "child_amount",
    "child_ages",
    "min_nights",
    "notes",
)

# Rows the template ships as worked examples. They are meant to be deleted; if
# they survive, skipping them beats importing a property called
# "EXAMPLE Temple Point - delete these rows".
EXAMPLE_MARKER = "EXAMPLE"


class Row(dict[str, Any]):
    """One sheet row, plus where it came from so an error can name it."""

    line: int = 0


def _rows_from_xlsx(path: Path) -> list[list[Any]]:
    # read_only streams the sheet rather than building the whole object graph,
    # which matters at a few thousand rows; data_only takes computed values
    # rather than formulae, because a rate typed as "=1200*1.16" should import
    # as the number an agent saw.
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _rows_from_csv(path: Path) -> list[list[Any]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [list(row) for row in csv.reader(handle)]


def read_sheet(path: str | Path) -> tuple[list[Row], list[str]]:
    """Return ``(rows, skipped)`` — the data rows, and notes about what was left out."""
    path = Path(path)
    if not path.is_file():
        raise AppError(f"No such intake sheet: {path}")
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        raw = _rows_from_xlsx(path)
    elif path.suffix.lower() == ".csv":
        raw = _rows_from_csv(path)
    else:
        raise AppError(
            f"{path.suffix} is not a readable intake sheet. Save it as .xlsx or .csv."
        )
    if not raw:
        raise AppError("That sheet is empty.")

    header = [clean(cell).lower() for cell in raw[0]]
    missing = [column for column in COLUMNS if column not in header]
    if missing:
        raise AppError(
            "The sheet is missing these columns: "
            + ", ".join(missing)
            + ". Start from docs/templates/rate-intake/hotel-rates.csv."
        )
    index = {column: header.index(column) for column in COLUMNS}

    rows: list[Row] = []
    skipped: list[str] = []
    for line, cells in enumerate(raw[1:], start=2):
        row = Row({name: cells[i] if i < len(cells) else None
                   for name, i in index.items()})
        row.line = line
        if not any(clean(value) for value in row.values()):
            continue
        if EXAMPLE_MARKER in clean(row["property_name"]).upper():
            skipped.append(f"row {line}: template example row")
            continue
        rows.append(row)
    return rows, skipped
